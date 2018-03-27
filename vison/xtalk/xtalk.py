#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""

Module for analysis of Cross-Talk data obtained with ROE-Tab injection.

Created on Thu Sep 14 16:11:09 2017

:author: Ruyman Azzollini
:contact: r.azzollini_at_ucl.ac.uk

"""

# IMPORT STUFF
from pdb import set_trace as stop
import numpy as np
from glob import glob
import os
import copy
import datetime
#import string as st

from scipy.optimize import curve_fit

from matplotlib import pyplot as plt

from vison.datamodel import ccd as ccdmod
from vison import __version__
from scipy import signal, stats
from astropy.io import fits as fts
# END IMPORT


isthere = os.path.exists


def f_fitXT(source,*p):

    fwc = 2.**16 # just a scaling factor
    victim = np.zeros_like(source)
    for ip in range(len(p)):
        victim += p[ip]*(source/fwc)**(ip+1)
    
    return fwc * victim

def sub_bgd(img,colend,VSTART=0,VEND=2086):
    """ """
    
    NX,NY = img.shape
    
    midbgd1d = np.nanmean(img[colend+20:-20,:],axis=0)
    midbgd2d = np.repeat(midbgd1d.reshape(1,NY),NX,axis=0)
    
    img = img - midbgd2d
    
    return img
    

def plot_quad(ax,Xtalk,Q):
        
   #Xtalk[Q] = dict(x=X,y=y,coefs=fitcoefs,model=model_ransac,
   #            inliers=inlier_mask,outliers=outlier_mask)
   x = Xtalk[Q]['x'].flatten().copy()
   ex = Xtalk[Q]['ex'].flatten().copy()
   y = Xtalk[Q]['y'].flatten().copy()
   ey = Xtalk[Q]['ey'].flatten().copy()
   
   coefs = Xtalk[Q]['coefs']
   
   yfit = f_fitXT(x,*coefs)
   
   xtalk_sig = np.nanstd(y-yfit)
   
   line_x = np.linspace(0.,6.5536E4,200)
   line_y = f_fitXT(line_x,*coefs)
   
   xtalk_val = Xtalk[Q]['xtalk'] * 2.**16
   xtalk_sig = Xtalk[Q]['std_xtalk'] * 2.**16
   
   ax.errorbar(x,y,xerr=ex,yerr=ey,fmt='o',color='b')
   if xtalk_val < 6.E4:
       label = r'$xtalk_{max}=%.2f \pm %.2f$ ADU' % (xtalk_val,xtalk_sig)
       ax.plot(line_x,line_y,'r--',label=label)
   else:
       ax.plot(line_x,line_y,'k-')
   plt.legend(loc='lower right',fontsize=8)
   ax.set_xlim((0,6.E4))
   if Q in ['G','H']: ax.set_xlabel('source [ADU]')
   if Q in ['E','H']: ax.set_ylabel('victim [ADU]')
   ax.ticklabel_format(style='sci',axis='x',scilimits=(0,0))
   
   ax.set_title(Q)
   
   return ax


def plotXtalk(Xtalk,suptitle='',figname=''):
    
    #ylim = [-5,20]
    
    fig = plt.figure(figsize=(8,6))
    ax1 = fig.add_subplot(221)
    ax1 = plot_quad(ax1,Xtalk,'E')
    #ax1.set_ylim(ylim)
    
    ax2 = fig.add_subplot(222)
    ax2 = plot_quad(ax2,Xtalk,'F')
    #ax2.set_ylim(ylim)
    
    ax3 = fig.add_subplot(223)
    ax3 = plot_quad(ax3,Xtalk,'H')
    #ax3.set_ylim(ylim)
    
    ax4 = fig.add_subplot(224)
    ax4 = plot_quad(ax4,Xtalk,'G')
    #ax4.set_ylim(ylim)
    
    plt.suptitle(suptitle)    
    plt.ticklabel_format(style='sci',axis='x',scilimits=(0,0))
    plt.ticklabel_format(style='sci',axis='y',scilimits=(0,0))
    plt.tight_layout()
    
    if figname != '': 
        plt.savefig(figname)
    else: 
        plt.show()
    plt.close()



def xtalk_fit(x,y):
        
    fitcoefs,pcov = curve_fit(f_fitXT,x,y,p0=[1.,0.])
    
    x = x.flatten().copy()
    y = y.flatten().copy()
    
    xsource = np.linspace(0,2.**16,200)
    ymodel = f_fitXT(xsource,*fitcoefs)
    
    ixmax = np.argmax(np.abs(ymodel))
    xtalk = ymodel[ixmax] / 2.**16
    
    std_xtalk = np.nanstd(y-f_fitXT(x,*fitcoefs)) / 2.**16
    
    
    xtalk_atmaxinj = y[np.argmax(x)]/x[np.argmax(x)]
    
    results = dict(x=x.copy(),y=y.copy(),coefs=fitcoefs,
                   xtalk=xtalk,std_xtalk=std_xtalk,
                   xtalk_atmaxinj=xtalk_atmaxinj)
    
    return results


def find_levels(img,colstart=1,colend=1600,rowstart=1,rowend=2086):
    
    histo = np.histogram(img[colstart:colend,rowstart:rowend],bins=200)
    
    bins = histo[1]
    nperbin = histo[0]
    
    nperbin[np.where(nperbin < 1E3)] = 0
    
    pnperbin = np.zeros(len(nperbin)+2)
    pnperbin[1:-1] = nperbin
    
    argrelmax=signal.argrelmax
        
    ixmax = argrelmax(pnperbin)[0]
    ixmax = [item - 1 for item in ixmax]
    
    levels = []
    npix = []
    
    for ix in ixmax:
        
        lowbound = bins[ix]
        hibound = bins[ix+1]
        
        ixsel = np.where((img >= lowbound) & (img <= hibound))
        val = np.nanmedian(img[ixsel])
      
        if ~np.isnan(val): levels.append(val)
        else: stop()
        
        npix.append(len(ixsel[0]))
    
        
    if np.any(np.isnan(levels)):
        raise RuntimeError
    
    return levels,npix



def processXtalk_single(CCDref,Qref,OBSID,thresholdinj=0.,colstart=1,colend=1600,
                        rowstart=1,rowend=2086,savefigs=False,log=None,datapath='',respath=''):
    """ 
    # Flow:
    #
    #   load ref-CCD, put readout node in lower-left corner (0,0) for every Q
    #   subtract bias (non-injected area)
    #
    #   find bin levels of (Q-CCD)ref
    #
    #   FOR CCD in CCDs:
    #
    #       load CCD
    #
    #       FOR Q in Quads != Qref:
    #
    #           put readout node of Q in lower-left corner (0,0)
    #           subtract bias (pre-scan)
    #   
    #           divide Q of CCD by Qref of CCDref.
    #
    #           for each bin:
    #               select values of (Q-CCD)ref within bin, mask
    #               measure average value of (Q-CCD) within mask
    # 
    #           plot av(Q-CCD)/av(Q-CCD)ref vs. av(bin)
    #           save slope (X-TALK) to log
    #
    """
    
    CCDs = [1,2,3]
    Quads = ['E','F','G','H']
    
    msg = 'Reference = CCD%i-Q:%s' % (CCDref,Qref)
    print msg
    
    if log is not None: log.info(msg)
    
    tCCDfileref = os.path.join(datapath,'EUC_%i_*_ROE1_CCD%i.fits' % (OBSID,CCDref,))
    
    CCDfileref = glob(tCCDfileref)[-1]
    
    ccdref = ccdmod.CCD(CCDfileref)
    imgref = ccdref.get_quad(Qref,canonical=True).copy()
    imgref = sub_bgd(imgref,colend)
    
    try: levels,nperlevel = find_levels(imgref,colstart,colend,rowstart,rowend)
    except RuntimeError: 
        log.info('Not Possible to analyze CCD%i,Q=%s' % (CCDref,Qref)) 
        return None
    
    
    mask = np.zeros_like(imgref)
    
    NX,NY = mask.shape
    
    Y,X = np.meshgrid(np.arange(NY),np.arange(NX),indexing='xy')
    
    refpix = []
    erefpix = []
    
    for level in levels:
        
        lowbound = min(level * 0.99,level-100)
        hibound = max(level * 1.01,level+100)

        ixsel = np.where((imgref >= lowbound) & (imgref <= hibound) & 
                         (X >= colstart) & (X<colend) &
                         (Y >= rowstart) & (Y<rowend))    
        mask[ixsel] = level
        pixels = imgref[ixsel]
        #refpix.append(np.nanmedian(pixels))
        
        vals = stats.sigmaclip(pixels,4,4).clipped
        nvals = len(vals)
        
        refpix.append(np.nanmean(vals))
        erefpix.append(np.nanstd(vals)/np.sqrt(nvals))
        
        #npixels = len(pixels)
        #refpix.append(np.nanmean(pixels))
        #erefpix.append(np.nanstd(pixels)/np.sqrt(npixels))

    
    refpix = np.array(refpix)
    erefpix = np.array(erefpix)
    
    #refpix -= refpix[0]
    
    maskfits = os.path.join(respath,'mask_%s%s.fits' % (CCDref,Qref))
    fts.writeto(maskfits,mask.astype('float32').transpose(),overwrite=True)
    
    Xtalk = {}
    
    for CCD in CCDs:
        
        print CCD
        
        tCCDfile = os.path.join(datapath,'EUC_%i_*_CCD%i.fits' % (OBSID,CCD,))
        CCDfile = glob(tCCDfile)[-1]
        
        ccd = ccdmod.CCD(CCDfile)
        
        Xtalk['CCD%i' % CCD] = dict()
                
        for Quad in Quads:

            #if not ((CCD == 3) and (Quad == 'F')): continue # TESTS   
            
            iimg = ccd.get_quad(Quad,canonical=True).copy()
            
            
            iimg = sub_bgd(iimg,colend)
            
            #iimg -= biaslevels['CCD%i' % CCD][Quad]
            
            xtalkpix = []
            extalkpix = []
            
            for level in levels:
                
                ixsel = np.where(mask == level)
                pixels = iimg[ixsel]
                
                #xtalkpix.append(np.nanmean(pixels)
                
                vals = stats.sigmaclip(pixels,4,4).clipped
                nvals = len(vals)
                
                xtalkpix.append(np.nanmean(vals))
                extalkpix.append(np.nanstd(vals)/np.sqrt(nvals))
            
            xtalkpix = np.array(xtalkpix)
            extalkpix = np.array(extalkpix)
            
            
            sellevels = np.where(refpix > thresholdinj)
            
           
            try: 
                results = xtalk_fit(refpix[sellevels],xtalkpix[sellevels])            
                log.info('CCD%i-%s, xtalk=%.2e' % (CCD,Quad,results['coefs'][0]))
            except: 
                log.info('CCD%i-%s, xtalk=?? (COULD NOT FIT)')
            
            results['ex'] = erefpix[sellevels].copy()
            results['ey'] = extalkpix[sellevels].copy()
            
            Xtalk['CCD%i' % CCD][Quad] = copy.deepcopy(results)
    
        if savefigs:
            figname = os.path.join(respath,'CCD%i_XTALK_ref%i%s.png' % \
                (CCD,CCDref,Qref))
        else: 
            figname=''
 
        suptitle='XTALK, CCD%i, source %i%s' % (CCD,CCDref,Qref)
        
        try: plotXtalk(Xtalk['CCD%i' % CCD],suptitle,figname)
        except KeyError: 
            pass
        
    return Xtalk
    

def PlotSummaryFig(Xtalk,suptitle,figname='',scale='RATIO'):
    """ """
    
    CCDs = [1,2,3]
    Quads = ['E','F','G','H']
    
    req = 0.15
    #logreq = np.log10(req/2.**16)
    
    fig = plt.figure(figsize=(8,7))
    ax = fig.add_subplot(111)
    
    for iCr,CCDref in enumerate(CCDs):
        for iQr,Qref in enumerate(Quads):
            
            for iC,CCD in enumerate(CCDs):
                for iQ,Q in enumerate(Quads):
                    CCDreftag = 'CCD%i' % CCDref
                    CCDtag = 'CCD%i' % CCD
                    
                    try: 
                        xtalk_dict = Xtalk[CCDreftag][Qref][CCDtag][Q]
                        
                        if scale == 'RATIO':
                        
                            ixtalk = xtalk_dict['xtalk'] #[0][0]
                            iextalk = xtalk_dict['std_xtalk']

                        elif scale == 'ADU':
                            
                            coefs = xtalk_dict['coefs']
                            xsource = np.linspace(0,2.**16,200)
                            yvictim = f_fitXT(xsource,*coefs)
                            
                            ixmax = np.argmax(np.abs(yvictim))
                            ixtalk = yvictim[ixmax]                   
                            iextalk = xtalk_dict['std_xtalk'] * 2.**16
                            
                    except:
                        
                        ixtalk = 1.E3
                        iextalk = 0.
                    
                    
                    y = iCr * 4 + iQr # source
                    x = iC * 4 + iQ # victim
                    
                    
                    eratio = np.abs(iextalk/ixtalk)
                    alpha = max(1.-eratio,0.1)
                    
                    if ixtalk > 0: color = 'b'
                    elif ixtalk < 0: color = 'r'
                    
                    #if (CCDref == 2) and (Qref == 'E') and (CCD==2) and (Q=='E') : stop()# TEST
                    
                    if scale == 'RATIO':
                    
                        if np.abs(ixtalk) < 0.9:
                            
                            ms = (np.abs(ixtalk) / req * 2.**16)*5.
                            #if ms <=0: color='g'
                            ms = max(ms,5)
                            
                            ax.scatter(x,y,color=color,s=[ms],alpha=alpha)
                            
                        elif (iCr==iC) and (iQ==iQr):
                            ax.scatter(x,y,marker='x',color='k') # self-cross-talk
                        elif np.isclose(ixtalk,1.E3): # empty value
                            ax.scatter(x,y,marker='s',facecolor='none',color='k')
                        else:
                            stop()
                        
                    elif scale == 'ADU':
                                                
                        if np.abs(ixtalk) < 500.:
                            
                            ms = (np.abs(ixtalk) / req)*5.
                            #if ms <=0: color='g'
                            ms = max(ms,5)
                            
                            ax.scatter(x,y,color=color,s=[ms],alpha=alpha)
                            
                        elif (iC==iCr) and (iQ==iQr):
                            ax.scatter(x,y,marker='x',color='k') # self-cross-talk
                        elif np.isclose(ixtalk,1.E3): # empty value
                            ax.scatter(x,y,marker='s',facecolor='none',color='k')
                        else:
                            stop()
                   
#                        # EM3 without modifications
#                        if (CCDref==2) and (Qref == 'H') and (CCD==2) and (Q=='G'):
#                            ms = (29.5/ req)*5.
#                            ms = max(ms,5)
#                            ax.scatter(x,y,facecolor='none',color='k',s=[ms])
#                        # EM3 without modifications
#                        if (CCDref==2) and (Qref == 'H') and (CCD==2) and (Q=='F'):
#                            ms = (3.6/ req)*5.
#                            ms = max(ms,5)
#                            ax.scatter(x,y,facecolor='none',color='k',s=[ms])

                        
    
    xaxisnames = ['1E','1F','1G','1H','2E','2F','2G','2H','3E','3F','3G','3H']    
    
    if scale == 'RATIO':
        
        ax.scatter(200,200,s=5.*(3E-4 / req * 2.**16),label='3E-4',color='k')
        ax.scatter(200,200,s=5.*(2E-4 / req * 2.**16),label='2E-4',color='k')
        ax.scatter(200,200,s=5.*(1E-4 / req * 2.**16),label='1E-4',color='k')
        ax.scatter(200,200,s=5.*(5E-5 / req * 2.**16),label='5E-5',color='k')
        ax.scatter(200,200,s=5.*(1E-5 / req * 2.**16),label='1E-5',color='k')
        ax.scatter(200,200,s=5.*(5.E-6 / req * 2.**16),label='5E-6',color='k')
        ax.scatter(200,200,s=5,label='<=2.3e-6',color='k')
    
    elif scale == 'ADU':
        
        ax.scatter(200,200,s=5.*(30./req),label='30 ADU',color='k')
        ax.scatter(200,200,s=5.*(20./req),label='20 ADU',color='k')
        ax.scatter(200,200,s=5.*(10./req),label='10 ADU',color='k')
        ax.scatter(200,200,s=5.*(5./req),label='5 ADU',color='k')
        ax.scatter(200,200,s=5.*(1./req),label='1 ADU',color='k')
        ax.scatter(200,200,s=5.*(0.5/req),label='0.5 ADU',color='k')
        ax.scatter(200,200,s=5.*(0.15/req),label='<=0.15 ADU',color='k')
        
    
    
    ax.set_xlim([-1,12])
    ax.set_ylim([-1,12])
    
    plt.xticks(range(len(xaxisnames)),xaxisnames,size='small')
    plt.yticks(range(len(xaxisnames)),xaxisnames,size='small')
    
    box = ax.get_position()
    ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])
    handles, labels = ax.get_legend_handles_labels()

    # Put a legend to the right of the current axis
    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
  
    
    ax.set_ylabel('Source Channel')
    ax.set_xlabel('Victim Channel')
    ax.set_title('Cross-talk - %s' % scale)
    
    
    if figname != '':        
        plt.savefig(figname)    
    else:
        plt.show()

    plt.close()
    

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
#from openpyxl.styles import PatternFill
import openpyxl


color_codes = dict(gray='B2A8A6',
                   red='F73910',
                   orange='F5AA08',
                   yellow='F5EA08',
                   blue='08B1F5',
                   green='08F510')

color_fills = dict()
for colkey in color_codes.keys():
    _color = openpyxl.styles.colors.Color(rgb=color_codes[colkey])
    color_fills[colkey] = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=_color)

class Report_Xtalk():
    """ """
    
    color_fills = color_fills
    
    threshold_red_ADU = 30. # ADU
    threshold_orange_ADU = 5. # ADU
    threshold_yellow_ADU = 2. # ADU
    
    
    def __init__(self,Xtalk):
        """ """
        
        self.CCDs = [1,2,3]
        self.Quads = ['E','F','G','H']
        
        self.data = Xtalk
        
        self.wb = Workbook()
        self.wb.create_sheet('Header',0)
        self.wb.create_sheet('Summary',1)
        self.wb.create_sheet('Xtalk_Ratio',2)
        self.wb.create_sheet('Xtalk_ADU',3)
        self.wb.create_sheet('Figures',4)
        
        std = self.wb.get_sheet_by_name('Sheet')
        self.wb.remove_sheet(std)
        
    
    def fill_Header(self):
        """ """
        
        headerdict = self.data['meta']
        
        headkeys = ['Date','ROE','ROE_FW','Injector','Injector_FW','Label',
        'Analysis_Date','vison']
        
        
        headerdict.update(dict(vison=__version__))
        
        self.wb['Header']['A1'] = 'XTALK Report'
        
        ix0 = 5
        for ik,key in enumerate(headkeys):
            self.wb['Header']['A%i' % (ix0+ik)] = key
            self.wb['Header']['B%i' % (ix0+ik)] = headerdict[key]
        
        self.adjust_columns_width(sheet='Header')
    
    def fill_Summary(self):
        """ """
        
        # STATS
    
        xtalks = []
        extalks = []
        couples = []
    
        for iCr,CCDref in enumerate(self.CCDs):
            for iQr,Qref in enumerate(self.Quads):
                
                CCDreftag = 'CCD%i' % CCDref
                
                for iCCD,CCD in enumerate(self.CCDs):
                    for iQ,Q in enumerate(self.Quads):
                        
                        if (iCCD == iCr) and (iQr == iQ): continue
                        
                        CCDtag = 'CCD%i' % CCD
                        
                        xtalk_dict = self.data[CCDreftag][Qref][CCDtag][Q]
                            
                        coefs = copy.deepcopy(xtalk_dict['coefs'])
                            
                        xsource = np.linspace(0,2.**16,200)
                        yvictim = f_fitXT(xsource,*coefs)
                                
                        ixmax = np.argmax(np.abs(yvictim))
                        ixtalk = yvictim[ixmax]                   
                        iextalk = xtalk_dict['std_xtalk'] * 2.**16
                        
                        xtalks.append(ixtalk)
                        extalks.append(iextalk)
                        
                        couples.append('%i%s-%i%s' % (iCr+1,Qref,iCCD+1,Q))
    
        xtalks = np.array(xtalks)
        axtalks = np.abs(xtalks)
        
        ncouples = len(couples)
        n05 = len(np.where(axtalks<=0.5)[0])
        n05_1 = len(np.where((axtalks>0.5) & (axtalks <=1.))[0])
        n1_2 = len(np.where((axtalks>1.0) & (axtalks <=2.))[0])
        n2_5 = len(np.where((axtalks>2.0) & (axtalks <=5.))[0])
        n5_100 = len(np.where((axtalks>5.) & (axtalks <=100.))[0])    
        ix_max = np.abs(xtalks).argmax()
        max_xtalk = xtalks[ix_max]
        max_couple = couples[ix_max]
        
        def assign_item_value(Item,Value,row):
            self.wb['Summary']['A%i' % row] = Item
            self.wb['Summary']['B%i' % row] = Value

        assign_item_value('# channels with max(x-talk) <= 0.5 ADU','%i/%i' % (n05,ncouples,),1)
        assign_item_value('# channels with 0.5 < max(x-talk) <= 1.0 ADU','%i/%i' % (n05_1,ncouples,),2)
        assign_item_value('# channels with 1.0 < max(x-talk) <= 2.0 ADU','%i/%i' % (n1_2,ncouples,),3)
        assign_item_value('# channels with 2.0 < max(x-talk) <= 5.0 ADU','%i/%i' % (n2_5,ncouples,),4)
        assign_item_value('# channels with 5.0 < max(x-talk) <= 100.0 ADU','%i/%i' % (n5_100,ncouples,),5)
        assign_item_value('Worst Cross-Talk [ADU]','%.2f' % max_xtalk,6)
        assign_item_value('Worst Cross-Talk is for S-V',max_couple,7)

               
        self.adjust_columns_width(sheet='Summary')
    
    def fill_Xtalk_Ratios(self):
        """ """

        ws = self.wb['Xtalk_Ratio']
        
        colnames = ['']
        for CCD in self.CCDs:
            for Q in self.Quads:
                colnames.append('%i%s' % (CCD,Q))
        
        for i,col in enumerate(ws.iter_cols(min_row=1,max_row=1,min_col=2,max_col=(len(colnames)-1)*3+1)):            
            if (i % 3) ==1:
                col[0].value = colnames[1+i/3]
            Q =colnames[1+i/3][-1] 
            if Q in ['E','G']:
                col[0].fill = self.color_fills['blue']
            elif Q in ['F','H']:
                col[0].fill = self.color_fills['green']
        
        subtitles = ['ratio','eratio','r_at_maxinj']        
        for i,col in enumerate(ws.iter_cols(min_row=2,max_row=2,min_col=2,max_col=(len(colnames)-1)*3+1)):
            col[0].value = subtitles[i%3]
        
        # Labelling reference channels
        
        channels = []
        for CCD in self.CCDs:
            for Q in self.Quads:
                channels.append('%i%s' % (CCD,Q))
        
        for i,row in enumerate(ws.iter_rows(min_row=3,max_row=3+len(channels)-1,min_col=1,max_col=1)):
            row[0].value = channels[i]
        
        for iCr,CCDref in enumerate(self.CCDs):
            for iQr,Qref in enumerate(self.Quads):
                
                irow = iCr*4+iQr+1+1+1
                
                for jC,CCD in enumerate(self.CCDs):
                    for jQ,Q, in enumerate(self.Quads):
                       
                        CCDreftag = 'CCD%i' % CCDref
                        CCDtag = 'CCD%i' % CCD
                        
                        xtalk_dict = self.data[CCDreftag][Qref][CCDtag][Q]
                        
                        ixtalk = xtalk_dict['xtalk']
                        iextalk = xtalk_dict['std_xtalk']
                        xtalk_atmaxinj = xtalk_dict['xtalk_atmaxinj']
                        
                        values = [ixtalk,iextalk,xtalk_atmaxinj]
                        
                        for k in range(3):
                            jcol = jC*4*3+jQ*3+k+1+1
                            jcolLetter = get_column_letter(jcol)
                        
                            ws['%s%i' % (jcolLetter,irow)] = values[k]
                            ws['%s%i' % (jcolLetter,irow)].number_format = '0.0E+00'
                            
                            if values[0] == 1.:
                                ws['%s%i' % (jcolLetter,irow)].fill = self.color_fills['gray']
                            
        
        
        self.adjust_columns_width(sheet='Xtalk_Ratio',minwidth=5)

    def fill_Xtalk_ADUs(self):
        """ """

        ws = self.wb['Xtalk_ADU']
        
        colnames = ['']
        for CCD in self.CCDs:
            for Q in self.Quads:
                colnames.append('%i%s' % (CCD,Q))
        
        for i,col in enumerate(ws.iter_cols(min_row=1,max_row=1,min_col=2,max_col=(len(colnames)-1)*2+1)):            
            if (i % 2) ==0:
                try: col[0].value = colnames[1+i/2]     
                except: stop()
            Q =colnames[1+i/2][-1] 
            if Q in ['E','G']:
                col[0].fill = self.color_fills['blue']
            elif Q in ['F','H']:
                col[0].fill = self.color_fills['green']
        
        subtitles = ['xtalk_lin','xtalk_at_maxinj']        
        for i,col in enumerate(ws.iter_cols(min_row=2,max_row=2,min_col=2,max_col=(len(colnames)-1)*2+1)):
            col[0].value = subtitles[i%2]
        
        # Labelling reference channels
        
        channels = []
        for CCD in self.CCDs:
            for Q in self.Quads:
                channels.append('%i%s' % (CCD,Q))
        
        for i,row in enumerate(ws.iter_rows(min_row=3,max_row=3+len(channels)-1,min_col=1,max_col=1)):
            row[0].value = channels[i]
        
        for iCr,CCDref in enumerate(self.CCDs):
            for iQr,Qref in enumerate(self.Quads):
                
                irow = iCr*4+iQr+1+1+1
                
                for jC,CCD in enumerate(self.CCDs):
                    for jQ,Q, in enumerate(self.Quads):
                       
                        CCDreftag = 'CCD%i' % CCDref
                        CCDtag = 'CCD%i' % CCD
                        
                        xtalk_dict = self.data[CCDreftag][Qref][CCDtag][Q]
                        
                        ixtalk = xtalk_dict['xtalk']
                        xtalk_atmaxinj = xtalk_dict['xtalk_atmaxinj']
                        
                        values = [ixtalk*2.**16,xtalk_atmaxinj*2.**16]
                            
                        
                        for k in range(2):
                            jcol = jC*4*2+jQ*2+k+1+1
                            jcolLetter = get_column_letter(jcol)
                        
                            ws['%s%i' % (jcolLetter,irow)] = values[k]
                            ws['%s%i' % (jcolLetter,irow)].number_format = '0.00'
                            
                            if ixtalk == 1.:
                                ws['%s%i' % (jcolLetter,irow)].fill = self.color_fills['gray']
                            else:
                                if np.any(np.array(values)>self.threshold_yellow_ADU):
                                    ws['%s%i' % (jcolLetter,irow)].fill = self.color_fills['yellow']
                                if np.any(np.array(values)>self.threshold_orange_ADU):
                                    ws['%s%i' % (jcolLetter,irow)].fill = self.color_fills['orange']
                                if np.any(np.array(values)>self.threshold_red_ADU):
                                    ws['%s%i' % (jcolLetter,irow)].fill = self.color_fills['red']
        
        self.adjust_columns_width(sheet='Xtalk_ADU',minwidth=5)
    
    def fill_Figures(self):
        """ """
        
        figsdict = self.data['figs']
        
        ws = self.wb['Figures']
        
        ws.add_image(openpyxl.drawing.image.Image(figsdict['RATIO']),'A1')
        ws.add_image(openpyxl.drawing.image.Image(figsdict['ADU']),'A40')
        
    
    
    def get_str_len(self,cell):
        if cell.value is None:
            return 0
        else:
            return len(str(cell.value))
        
    def adjust_columns_width(self,sheet,minwidth=0):
        """ """
        
        column_widths = []
        for row in self.wb[sheet]:
            for i, cell in enumerate(row):                
                if len(column_widths)>i:                    
                    if self.get_str_len(cell)>column_widths[i]:
                        column_widths[i] = self.get_str_len(cell)
                else:
                    column_widths += [self.get_str_len(cell)]

        for i, column_width in enumerate(column_widths):
            self.wb[sheet].column_dimensions[get_column_letter(i+1)].width = max([column_width,minwidth])
    
    def save(self,filename):
        self.wb.save(filename)
    
        
    
    
