#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Mar 26 12:07:54 2018

:author: raf
:contact: r.azzollini_at_ucl.ac.uk

"""

# IMPORT STUFF
from pdb import set_trace as stop
from openpyxl import Workbook
from openpyxl.drawing.image import Image
# END IMPORT

def test0():
    """Just a dummy test to show we can use openpyxl"""
    
    wb = Workbook()
    ws = wb.active    
    ws1 = wb.create_sheet('Mysheet')
    ws2 = wb.create_sheet('Mysheet',0)
    ws1.title= "New Title"    
    ws.sheet_properties.tabColor = "1072BA"    
    ws3 = wb["New Title"]    
    ws['A4'] = 4    
    
    img = Image('../data/Eyegore.gif')
    stop()
    
    ws2.add_image(img,'A1')
    
    print (wb.sheetnames)
    
    
    wb.save('Test0.xlsx')    

    
if __name__ == '__main__':
    test0()
    