#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Mar 26 12:07:54 2018

:author: Ruyman Azzollini
:contact: r.azzollini_at_ucl.ac.uk

"""

# IMPORT STUFF
from pdb import set_trace as stop
import openpyxl as oxl
from openpyxl.utils import dataframe as odataframe
import pandas
import copy
from collections import OrderedDict
# END IMPORT

color_codes = dict(gray='B2A8A6',
                   red='F73910',
                   orange='F5AA08',
                   yellow='F5EA08',
                   blue='08B1F5',
                   green='08F510')

color_fills = dict()
for colkey in color_codes.keys():
    _color = oxl.styles.colors.Color(rgb=color_codes[colkey])
    color_fills[colkey] = oxl.styles.fills.PatternFill(
        patternType='solid', fgColor=_color)


def get_str_len(cell):
    if cell.value is None:
        return 0
    else:
        return len(str(cell.value))


class ReportXL(object):

    color_fills = color_fills

    def __init__(self, datadict=None):

        self.data = copy.deepcopy(datadict)

        self.wb = oxl.Workbook()
        std = self.wb['Sheet']
        self.wb.remove(std)

    def adjust_columns_width(self, sheetname, minwidth=0):
        """ """

        column_widths = []
        for row in self.wb[sheetname]:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if get_str_len(cell) > column_widths[i]:
                        column_widths[i] = get_str_len(cell)
                else:
                    column_widths += [get_str_len(cell)]

        for i, column_width in enumerate(column_widths):
            self.wb[sheetname].column_dimensions[oxl.utils.get_column_letter(
                i+1)].width = max([column_width, minwidth])

    def dict_to_sheet(self, indict, sheetname, keys=None, title=''):
        """ """

        if keys is None:
            keys = indict.keys()

        printdict = OrderedDict()
        for key in keys:
            printdict[key] = str(indict[key].__repr__())

        self.wb[sheetname]['A1'] = title

        ix0 = 5
        for ik, key in enumerate(keys):
            self.wb[sheetname]['A%i' % (ix0+ik)] = key
            self.wb[sheetname]['B%i' % (ix0+ik)] = printdict[key]

        self.adjust_columns_width(sheetname=sheetname)

    def df_to_sheet(self, df, sheetname, minwidth=10, index=False, header=True):
        """ """

        ws = self.wb[sheetname]
        df_to_rows = odataframe.dataframe_to_rows

        for r in df_to_rows(df, index=index, header=header):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        self.adjust_columns_width(sheetname, minwidth=minwidth)

    def add_image(self, figpath, sheetname, cell='A1'):
        """ """
        self.wb[sheetname].add_image(oxl.drawing.image.Image(figpath), cell)

    def save(self, filename):
        self.wb.save(filename)


def test0():
    """Just a dummy test to show we can use openpyxl"""

    wb = oxl.Workbook()
    ws = wb.active
    ws1 = wb.create_sheet('Mysheet')
    ws2 = wb.create_sheet('Mysheet', 0)
    ws1.title = "New Title"
    ws.sheet_properties.tabColor = "1072BA"
    #ws3 = wb["New Title"]
    ws['A4'] = 4

    img = oxl.drawing.image.Image('../data/Eyegore.gif')

    ws2.add_image(img, 'A1')

    print (wb.sheetnames)

    wb.save('Test0.xlsx')


if __name__ == '__main__':
    test0()
